"""
CEBRAC — Gerador de Relatórios de Presença
Gera relatórios em Excel (.xlsx) e PDF por turma.

Uso:
  python relatorio.py --turma <id_ou_nome> --formato excel
  python relatorio.py --turma <id_ou_nome> --formato pdf
  python relatorio.py --todas --formato excel
  python relatorio.py --todas --formato ambos
"""

import argparse
import os
import sys
import requests
from datetime import datetime

# ─── CONFIG ───────────────────────────────────────────────
SUPABASE_URL = 'https://joszmqohhceuxhsjxxcr.supabase.co'
SUPABASE_KEY = 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Impvc3ptcW9oaGNldXhoc2p4eGNyIiwicm9sZSI6ImFub24iLCJpYXQiOjE3ODEyODEwNjEsImV4cCI6MjA5Njg1NzA2MX0.sSPFSYVtNGbgelrcQNK2mS-1KCk13A5ROid7E0YewIg'
HEADERS = {'apikey': SUPABASE_KEY, 'Authorization': f'Bearer {SUPABASE_KEY}'}
SAIDA = './relatorios'
os.makedirs(SAIDA, exist_ok=True)

# ─── SUPABASE ─────────────────────────────────────────────
def buscar(tabela, params=''):
    r = requests.get(f'{SUPABASE_URL}/rest/v1/{tabela}?{params}', headers=HEADERS)
    return r.json()

def buscar_turmas():
    return buscar('turmas', 'select=*&ativa=eq.true&order=nome')

def buscar_turma(turma_id):
    r = buscar('turmas', f'select=*&id=eq.{turma_id}')
    return r[0] if r else None

def buscar_alunos_turma(turma_id):
    return buscar('turma_alunos', f'select=*&turma_id=eq.{turma_id}&order=nome')

def buscar_chamadas(turma_id):
    return buscar('chamadas', f'select=*&turma_id=eq.{turma_id}&order=numero_aula')

def buscar_presencas_chamada(chamada_id):
    return buscar('chamada_presencas', f'select=*&chamada_id=eq.{chamada_id}')

def formatar_data(data_str):
    if not data_str:
        return '—'
    try:
        d = datetime.strptime(data_str[:10], '%Y-%m-%d')
        return d.strftime('%d/%m/%Y')
    except:
        return data_str[:10]

# ─── MONTAR DADOS ─────────────────────────────────────────
def montar_dados_turma(turma_id):
    turma = buscar_turma(turma_id)
    if not turma:
        print(f'Turma {turma_id} não encontrada.')
        return None

    alunos = buscar_alunos_turma(turma_id)
    chamadas = buscar_chamadas(turma_id)

    # Mapa: contrato -> {aula_n: status}
    mapa = {a['contrato']: {'nome': a['nome'], 'aulas': {}} for a in alunos}

    datas_aulas = {}
    for ch in chamadas:
        n = ch['numero_aula']
        datas_aulas[n] = formatar_data(ch.get('data_aula'))
        presencas = buscar_presencas_chamada(ch['id'])
        for p in presencas:
            c = p['contrato']
            if c not in mapa:
                mapa[c] = {'nome': p['nome'], 'aulas': {}}
            mapa[c]['aulas'][n] = p.get('status', '—')

    return {
        'turma': turma,
        'alunos': alunos,
        'mapa': mapa,
        'datas': datas_aulas,
        'chamadas': chamadas,
    }

# ─── EXCEL ────────────────────────────────────────────────
def gerar_excel(dados):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    turma = dados['turma']
    mapa = dados['mapa']
    datas = dados['datas']

    wb = Workbook()
    ws = wb.active
    ws.title = turma['nome'][:30]

    # Cores
    COR_HEADER    = 'E85D04'
    COR_PRESENTE  = 'D4EDDA'
    COR_FALTA     = 'F8D7DA'
    COR_TITULO    = 'FFF3ED'
    BRANCO        = 'FFFFFF'

    borda = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    # Título
    ws.merge_cells('A1:J1')
    ws['A1'] = f'CEBRAC — Relatório de Presença'
    ws['A1'].font = Font(bold=True, size=14, color=COR_HEADER)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:J2')
    horario = ''
    if turma.get('hora_inicio'):
        horario = f" · {turma['hora_inicio']}"
        if turma.get('hora_fim'):
            horario += f" – {turma['hora_fim']}"
    ws['A2'] = f"{turma['nome']} · Prof. {turma['professor_nome']}{horario}"
    ws['A2'].font = Font(size=11, color='555555')
    ws['A2'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A3:J3')
    ws['A3'] = f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A3'].font = Font(size=9, color='888888')
    ws['A3'].alignment = Alignment(horizontal='center')

    ws.append([])

    # Cabeçalho
    cabecalho = ['#', 'Nome', 'Contrato',
                 f"Aula 1\n{datas.get(1,'—')}",
                 f"Aula 2\n{datas.get(2,'—')}",
                 f"Aula 3\n{datas.get(3,'—')}",
                 f"Aula 4\n{datas.get(4,'—')}",
                 'Presenças', 'Faltas', 'Situação']

    ws.append(cabecalho)
    header_row = ws.max_row
    for col in range(1, 11):
        cell = ws.cell(row=header_row, column=col)
        cell.font = Font(bold=True, color=BRANCO, size=10)
        cell.fill = PatternFill('solid', fgColor=COR_HEADER)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = borda
    ws.row_dimensions[header_row].height = 36

    # Dados
    alunos_ordenados = sorted(mapa.items(), key=lambda x: x[1]['nome'])
    total_p = total_f = 0

    for i, (contrato, info) in enumerate(alunos_ordenados, 1):
        aulas = info['aulas']
        presencas = sum(1 for n in [1,2,3,4] if aulas.get(n) == 'C')
        faltas = sum(1 for n in [1,2,3,4] if aulas.get(n) == 'F')
        situacao = '✓ OK' if faltas == 0 else ('⚠ Atenção' if faltas <= 1 else '✗ Crítico')
        total_p += presencas
        total_f += faltas

        row = [
            i,
            info['nome'],
            contrato,
            aulas.get(1, '—'),
            aulas.get(2, '—'),
            aulas.get(3, '—'),
            aulas.get(4, '—'),
            presencas,
            faltas,
            situacao
        ]
        ws.append(row)
        r = ws.max_row
        fill_linha = PatternFill('solid', fgColor='F9F9F7' if i % 2 == 0 else BRANCO)

        for col in range(1, 11):
            cell = ws.cell(row=r, column=col)
            cell.border = borda
            cell.alignment = Alignment(horizontal='center', vertical='center')

            # Cor nas aulas
            if col in [4,5,6,7]:
                val = cell.value
                if val == 'C':
                    cell.fill = PatternFill('solid', fgColor=COR_PRESENTE)
                    cell.font = Font(bold=True, color='155724')
                elif val == 'F':
                    cell.fill = PatternFill('solid', fgColor=COR_FALTA)
                    cell.font = Font(bold=True, color='721C24')
                else:
                    cell.fill = fill_linha
            elif col == 10:
                if 'Crítico' in str(cell.value):
                    cell.fill = PatternFill('solid', fgColor=COR_FALTA)
                    cell.font = Font(bold=True, color='721C24')
                elif 'Atenção' in str(cell.value):
                    cell.fill = PatternFill('solid', fgColor='FFF3CD')
                    cell.font = Font(bold=True, color='856404')
                else:
                    cell.fill = PatternFill('solid', fgColor=COR_PRESENTE)
                    cell.font = Font(bold=True, color='155724')
            elif col == 2:
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.fill = fill_linha
            else:
                cell.fill = fill_linha

    # Totais
    ws.append([])
    total_row = ['', 'TOTAL', '', '', '', '', '', total_p, total_f, '']
    ws.append(total_row)
    r = ws.max_row
    for col in range(1, 11):
        cell = ws.cell(row=r, column=col)
        cell.font = Font(bold=True, size=10)
        cell.fill = PatternFill('solid', fgColor=COR_TITULO)
        cell.border = borda
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Larguras
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 36
    ws.column_dimensions['C'].width = 12
    for col in ['D','E','F','G']:
        ws.column_dimensions[col].width = 14
    ws.column_dimensions['H'].width = 11
    ws.column_dimensions['I'].width = 8
    ws.column_dimensions['J'].width = 12

    nome_arquivo = f"{SAIDA}/{turma['nome']}_presencas.xlsx"
    wb.save(nome_arquivo)
    print(f"  Excel salvo: {nome_arquivo}")
    return nome_arquivo

# ─── PDF ──────────────────────────────────────────────────
def gerar_pdf(dados):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    turma = dados['turma']
    mapa = dados['mapa']
    datas = dados['datas']

    nome_arquivo = f"{SAIDA}/{turma['nome']}_presencas.pdf"
    doc = SimpleDocTemplate(nome_arquivo, pagesize=landscape(A4),
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)

    styles = getSampleStyleSheet()
    LARANJA = colors.HexColor('#E85D04')
    VERDE   = colors.HexColor('#D4EDDA')
    VVERDE  = colors.HexColor('#155724')
    VERMELHO= colors.HexColor('#F8D7DA')
    VVERM   = colors.HexColor('#721C24')
    CINZA   = colors.HexColor('#F4F4F0')

    titulo_style = ParagraphStyle('titulo', parent=styles['Title'],
                                   fontSize=16, textColor=LARANJA, spaceAfter=4)
    sub_style = ParagraphStyle('sub', parent=styles['Normal'],
                                fontSize=10, textColor=colors.HexColor('#555555'), spaceAfter=2)
    data_style = ParagraphStyle('data', parent=styles['Normal'],
                                 fontSize=8, textColor=colors.HexColor('#888888'), spaceAfter=12)

    horario = ''
    if turma.get('hora_inicio'):
        horario = f" · {turma['hora_inicio']}"
        if turma.get('hora_fim'):
            horario += f" – {turma['hora_fim']}"

    story = [
        Paragraph('CEBRAC — Relatório de Presença', titulo_style),
        Paragraph(f"{turma['nome']} · Prof. {turma['professor_nome']}{horario}", sub_style),
        Paragraph(f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}", data_style),
    ]

    # Cabeçalho da tabela
    cab = ['#', 'Nome', 'Contrato',
           f"Aula 1\n{datas.get(1,'—')}",
           f"Aula 2\n{datas.get(2,'—')}",
           f"Aula 3\n{datas.get(3,'—')}",
           f"Aula 4\n{datas.get(4,'—')}",
           'P', 'F', 'Situação']

    alunos_ordenados = sorted(mapa.items(), key=lambda x: x[1]['nome'])
    table_data = [cab]
    style_cmds = [
        ('BACKGROUND', (0,0), (-1,0), LARANJA),
        ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
        ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0,0), (-1,0), 9),
        ('ALIGN',      (0,0), (-1,-1), 'CENTER'),
        ('VALIGN',     (0,0), (-1,-1), 'MIDDLE'),
        ('ALIGN',      (1,0), (1,-1), 'LEFT'),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F9F9F7')]),
        ('GRID',       (0,0), (-1,-1), 0.4, colors.HexColor('#CCCCCC')),
        ('FONTSIZE',   (0,1), (-1,-1), 8),
        ('TOPPADDING', (0,0), (-1,-1), 5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
    ]

    total_p = total_f = 0
    for i, (contrato, info) in enumerate(alunos_ordenados, 1):
        aulas = info['aulas']
        p = sum(1 for n in [1,2,3,4] if aulas.get(n) == 'C')
        f = sum(1 for n in [1,2,3,4] if aulas.get(n) == 'F')
        sit = '✓ OK' if f == 0 else ('⚠ Aten.' if f <= 1 else '✗ Crít.')
        total_p += p
        total_f += f
        row_data = [i, info['nome'], contrato,
                    aulas.get(1,'—'), aulas.get(2,'—'), aulas.get(3,'—'), aulas.get(4,'—'),
                    p, f, sit]
        table_data.append(row_data)
        r = len(table_data) - 1
        # Colorir aulas
        for col_idx, n in [(3,1),(4,2),(5,3),(6,4)]:
            v = aulas.get(n)
            if v == 'C':
                style_cmds.append(('BACKGROUND', (col_idx,r), (col_idx,r), VERDE))
                style_cmds.append(('TEXTCOLOR',  (col_idx,r), (col_idx,r), VVERDE))
                style_cmds.append(('FONTNAME',   (col_idx,r), (col_idx,r), 'Helvetica-Bold'))
            elif v == 'F':
                style_cmds.append(('BACKGROUND', (col_idx,r), (col_idx,r), VERMELHO))
                style_cmds.append(('TEXTCOLOR',  (col_idx,r), (col_idx,r), VVERM))
                style_cmds.append(('FONTNAME',   (col_idx,r), (col_idx,r), 'Helvetica-Bold'))
        # Colorir situação
        if '✗' in sit:
            style_cmds.append(('BACKGROUND', (9,r), (9,r), VERMELHO))
            style_cmds.append(('TEXTCOLOR',  (9,r), (9,r), VVERM))
            style_cmds.append(('FONTNAME',   (9,r), (9,r), 'Helvetica-Bold'))
        elif '⚠' in sit:
            style_cmds.append(('BACKGROUND', (9,r), (9,r), colors.HexColor('#FFF3CD')))
            style_cmds.append(('TEXTCOLOR',  (9,r), (9,r), colors.HexColor('#856404')))

    # Linha de totais
    table_data.append(['', 'TOTAL', '', '', '', '', '', total_p, total_f, ''])
    r_tot = len(table_data) - 1
    style_cmds += [
        ('BACKGROUND', (0,r_tot), (-1,r_tot), CINZA),
        ('FONTNAME',   (0,r_tot), (-1,r_tot), 'Helvetica-Bold'),
        ('FONTSIZE',   (0,r_tot), (-1,r_tot), 9),
    ]

    col_widths = [1*cm, 7.5*cm, 2.5*cm, 2.5*cm, 2.5*cm, 2.5*cm, 2.5*cm, 1.2*cm, 1.2*cm, 2.2*cm]
    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle(style_cmds))
    story.append(t)

    doc.build(story)
    print(f"  PDF salvo: {nome_arquivo}")
    return nome_arquivo

# ─── MAIN ─────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description='Gerador de relatórios CEBRAC')
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument('--turma', help='ID ou nome da turma')
    group.add_argument('--todas', action='store_true', help='Gerar relatório de todas as turmas')
    parser.add_argument('--formato', choices=['excel', 'pdf', 'ambos'], default='ambos')
    args = parser.parse_args()

    turmas = []
    if args.todas:
        turmas = buscar_turmas()
        print(f"Gerando relatórios para {len(turmas)} turmas...")
    else:
        # Buscar por ID ou nome
        todas = buscar_turmas()
        turmas = [t for t in todas if t['id'] == args.turma or args.turma.upper() in t['nome'].upper()]
        if not turmas:
            print(f"Turma '{args.turma}' não encontrada.")
            sys.exit(1)

    for turma in turmas:
        print(f"\nProcessando: {turma['nome']}")
        dados = montar_dados_turma(turma['id'])
        if not dados:
            continue
        if args.formato in ['excel', 'ambos']:
            gerar_excel(dados)
        if args.formato in ['pdf', 'ambos']:
            gerar_pdf(dados)

    print(f"\nConcluído! Arquivos salvos em: {SAIDA}/")

if __name__ == '__main__':
    main()
